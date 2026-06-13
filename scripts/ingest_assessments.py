#!/usr/bin/env python3
"""
ingest_assessments.py — parse the Balochistan "Digital Assessments" Excel tree
into a real, structured question bank + knowledge base the app loads and the AI
grounds on (RAG). Replaces all dummy/mock question data.

Handles four real-world layouts found in the tree:
  - std : header row with  question | option1 | correct1 | ...   (also "questions"/"option 1")
  - apr : answer-per-row    questionId | Answer/UrduAnswer | ... | correctAnswer(0/1)   (multi-sheet)
  - pos : headerless fixed positional layout (grade,subject,unit,topic,title,question,_,opt,flag,...)
Catalogue/metadata sheets (Package Code / SLO inventories) have no questions and are skipped.

Usage:  python scripts/ingest_assessments.py [ROOT] [OUT]
"""
import sys, os, json, re, glob

try:
    import openpyxl
except Exception:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = sys.argv[1] if len(sys.argv) > 1 else "docs/Digital Assessments"
OUT  = sys.argv[2] if len(sys.argv) > 2 else "src/assets/data/assessments"

def norm(s):
    return ("" if s is None else str(s)).strip()

def norm_key(s):
    return re.sub(r"\s+", "", norm(s).lower())

def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", norm(s).lower()).strip("-")

def grade_from_path(parts):
    for p in parts:
        m = re.match(r"grade\s*0*([0-9]+)", p, re.I)
        if m: return m.group(1)
        if p.strip().upper() in ("ECD", "ECD-5"): return "ECD"
    return "?"

KNOWN_SUBJECTS = ["Biology","Chemistry","Physics","Mathematics","Math","English","Urdu",
                  "Science","Pak Studies","Pakistan Studies","General Knowledge","Computer",
                  "Islamiat","Social Studies"]

def subject_from_path(parts):
    for p in parts:
        for s in KNOWN_SUBJECTS:
            if p.strip().lower() == s.lower():
                return "Mathematics" if s == "Math" else s
    return None

def unit_from_path(parts):
    for p in parts:
        if re.match(r"unit\s*", p, re.I):
            return p.strip()
    return ""

EMPTY = ("", "null", "-")

def _cell(row, idx):
    return norm(row[idx]) if (idx is not None and idx < len(row)) else ""

def detect_header(ws):
    """(kind, header_row, {normkey:col}) — kind in {'std','apr'} or (None,None,None)."""
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=14, values_only=True)):
        keys = [norm_key(c) for c in row]
        kset = set(keys)
        if ("question" in kset or "questions" in kset) and any(re.fullmatch(r"option\d+", k) for k in keys):
            return "std", ri, {k: i for i, k in enumerate(keys) if k}
        if "questionid" in kset and "correctanswer" in kset:
            return "apr", ri, {k: i for i, k in enumerate(keys) if k}
    return None, None, None

def parse_articles(wb):
    arts = {}
    if "Article" in wb.sheetnames:
        for row in wb["Article"].iter_rows(min_row=2, values_only=True):
            if not row: continue
            aid, txt = norm(row[0]), (norm(row[1]) if len(row) > 1 else "")
            if aid and txt: arts[aid] = txt
    return arts

def parse_std(ws, hr, cols, ctx, articles):
    out = []
    qcol = cols.get("question", cols.get("questions"))
    for row in ws.iter_rows(min_row=hr + 2, values_only=True):
        q = _cell(row, qcol)
        if not q: continue
        options, correct = [], -1
        for k in range(1, 7):
            oi = cols.get(f"option{k}")
            if oi is None: continue
            opt = _cell(row, oi)
            if opt == "": continue
            if _cell(row, cols.get(f"correct{k}")).upper().startswith("Y"):
                correct = len(options)
            options.append(opt)
        if len(options) < 2: continue
        if correct < 0: correct = 0
        aid = _cell(row, cols.get("articleid"))
        item = dict(ctx, q=q, options=options, correct=correct,
                    topic=_cell(row, cols.get("topic")) or ctx["unit"] or ctx["subject"],
                    skill=_cell(row, cols.get("skill", cols.get("skills"))),
                    feedback=_cell(row, cols.get("feedback")))
        if aid:
            item["articleId"] = aid
            if aid in articles: item["article"] = articles[aid]
        out.append(item)
    return out

def parse_apr(ws, hr, cols, ctx):
    """Answer-per-row: group by questionId; option text from Answer/UrduAnswer; correct via correctAnswer."""
    qi, qt, ans, urdu, corr = (cols.get("questionid"), cols.get("question"),
                               cols.get("answer"), cols.get("urduanswer"), cols.get("correctanswer"))
    groups, order = {}, []
    for row in ws.iter_rows(min_row=hr + 2, values_only=True):
        qid = _cell(row, qi)
        if not qid: continue
        en, ur = _cell(row, ans), _cell(row, urdu)
        opt = en if en.lower() not in EMPTY else ur
        if opt.lower() in EMPTY: opt = ur if ur.lower() not in EMPTY else en
        if opt.lower() in EMPTY: continue
        if qid not in groups:
            groups[qid] = {"q": _cell(row, qt), "options": [], "correct": -1}
            order.append(qid)
        g = groups[qid]
        if _cell(row, corr) in ("1", "Y", "y", "true", "True"):
            g["correct"] = len(g["options"])
        g["options"].append(opt)
    out = []
    for qid in order:
        g = groups[qid]
        if not g["q"] or len(g["options"]) < 2: continue
        if g["correct"] < 0: g["correct"] = 0
        out.append(dict(ctx, q=g["q"], options=g["options"], correct=g["correct"],
                        topic=ctx["unit"] or ctx["subject"], skill="", feedback=""))
    return out

def parse_positional(ws, ctx):
    """Headerless: grade,subject,unit,topic,title,question,(gap),opt1,flag1,opt2,flag2,..."""
    out = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 9: continue
        c = [norm(x) for x in row]
        q = c[5] if len(c) > 5 else ""
        if not q: continue
        options, correct, k = [], -1, 7
        while k + 1 < len(c):
            opt, flag = c[k], c[k + 1]
            if opt != "":
                if flag.upper().startswith("Y"):
                    correct = len(options)
                options.append(opt)
            k += 2
        if len(options) < 2 or correct < 0:   # require an explicit correct flag (avoid garbage)
            continue
        out.append(dict(ctx, q=q, options=options, correct=correct,
                        topic=(c[3] if len(c) > 3 else "") or ctx["unit"] or ctx["subject"],
                        skill=(c[4] if len(c) > 4 else ""), feedback=""))
    return out

def parse_file(path, rel_parts):
    ctx = {
        "grade": grade_from_path(rel_parts),
        "subject": subject_from_path(rel_parts) or "General",
        "unit": unit_from_path(rel_parts),
        "paper": os.path.splitext(os.path.basename(path))[0],
    }
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return [], f"open-fail: {e}"
    out, matched = [], False
    articles = parse_articles(wb)
    for ws in wb.worksheets:
        kind, hr, cols = detect_header(ws)
        if kind == "std":
            matched = True; out += parse_std(ws, hr, cols, ctx, articles)
        elif kind == "apr":
            matched = True; out += parse_apr(ws, hr, cols, ctx)
    if not matched:
        out += parse_positional(wb.worksheets[0], ctx)
        if not out:
            return [], "no-questions"
    return out, None

def main():
    files = sorted(glob.glob(os.path.join(ROOT, "**", "*.xlsx"), recursive=True))
    print(f"Found {len(files)} xlsx under {ROOT}")
    banks, index = {}, {}
    stats = {"files": len(files), "ok": 0, "skipped": [], "questions": 0}
    for n, path in enumerate(files):
        rel = os.path.relpath(path, ROOT).replace("\\", "/")
        items, err = parse_file(path, rel.split("/"))
        if err and not items:
            stats["skipped"].append({"file": rel, "reason": err}); continue
        stats["ok"] += 1
        for it in items:
            key = (it["grade"], it["subject"])
            banks.setdefault(key, []).append(it)
            g = index.setdefault(it["grade"], {})
            s = g.setdefault(it["subject"], {"units": set(), "topics": set(), "count": 0})
            if it["unit"]: s["units"].add(it["unit"])
            if it["topic"]: s["topics"].add(it["topic"])
            s["count"] += 1
        stats["questions"] += len(items)
        if (n + 1) % 200 == 0:
            print(f"  ...{n+1}/{len(files)} files, {stats['questions']} questions")

    os.makedirs(os.path.join(OUT, "bank"), exist_ok=True)
    catalogue = {}
    for grade, subs in index.items():
        catalogue[grade] = {}
        for subject, info in subs.items():
            fname = f"{slug(grade)}__{slug(subject)}.json"
            with open(os.path.join(OUT, "bank", fname), "w", encoding="utf-8") as f:
                json.dump(banks[(grade, subject)], f, ensure_ascii=False)
            catalogue[grade][subject] = {"file": f"bank/{fname}", "count": info["count"],
                                         "units": sorted(info["units"]), "topics": sorted(info["topics"])[:200]}
    with open(os.path.join(OUT, "index.json"), "w", encoding="utf-8") as f:
        json.dump(catalogue, f, ensure_ascii=False, indent=1)
    with open(os.path.join(OUT, "_stats.json"), "w", encoding="utf-8") as f:
        json.dump({"files": stats["files"], "ok": stats["ok"], "questions": stats["questions"],
                   "skipped_count": len(stats["skipped"]), "skipped_sample": stats["skipped"][:20]},
                  f, ensure_ascii=False, indent=1)
    print(f"DONE: {stats['ok']}/{len(files)} files parsed, {stats['questions']} questions, "
          f"{len(banks)} grade-subject banks. Skipped {len(stats['skipped'])}.")

if __name__ == "__main__":
    main()
